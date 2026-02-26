"""
Excel output writer using openpyxl.
Produces formatted .xlsx files with frozen header rows and auto-fitted column widths.

Functions:
    write_exhibitors_excel(records, path)   → output/exhibitors.xlsx
    write_team_members_excel(records, path) → output/team_members.xlsx
"""
import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger("expowest_scraper.output")

# ── Column definitions ─────────────────────────────────────────────────────────

EXHIBITOR_COLUMNS = [
    ("exhibitor_name",      "Exhibitor Name"),
    ("booth_number",        "Booth Number"),
    ("information",         "Description / Information"),
    ("product_categories",  "Product Categories"),
    ("hall",                "Hall"),
    ("country",             "Country"),
    ("company_url",         "Company URL"),
    ("social_media_links",  "Social Media Links"),
    ("source_url",          "Source URL"),
]

TEAM_COLUMNS = [
    ("exhibitor_name",      "Exhibitor Name"),
    ("team_member_name",    "Team Member Name"),
    ("job_title",           "Job Title"),
]

# Header style: bold white text on dark green background
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F5C2E", end_color="1F5C2E", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
DATA_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ── Public API ─────────────────────────────────────────────────────────────────

def write_exhibitors_excel(
    records: list[dict],
    path: Path = config.EXHIBITORS_OUTPUT,
) -> None:
    """Write exhibitor records to an .xlsx file."""
    _write_excel(records, EXHIBITOR_COLUMNS, path, sheet_name="Exhibitors")
    logger.info(f"Wrote {len(records)} exhibitors → {path}")


def write_team_members_excel(
    records: list[dict],
    path: Path = config.TEAM_MEMBERS_OUTPUT,
) -> None:
    """Write team member records to an .xlsx file."""
    _write_excel(records, TEAM_COLUMNS, path, sheet_name="Team Members")
    logger.info(f"Wrote {len(records)} team members → {path}")


# ── Internal helpers ───────────────────────────────────────────────────────────

def _write_excel(
    records: list[dict],
    columns: list[tuple[str, str]],
    path: Path,
    sheet_name: str = "Sheet1",
) -> None:
    """
    Write records to an Excel file.

    Args:
        records:    List of dicts; keys should match the field names in columns.
        columns:    Ordered list of (field_name, display_header) tuples.
        path:       Output file path (.xlsx).
        sheet_name: Name of the worksheet tab.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    field_names = [col[0] for col in columns]
    headers = [col[1] for col in columns]

    # ── Write header row ───────────────────────────────────────────────────────
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Write data rows ────────────────────────────────────────────────────────
    for record in records:
        row_values = [_safe_str(record.get(field, "")) for field in field_names]
        ws.append(row_values)
        row_idx = ws.max_row
        for col_idx in range(1, len(field_names) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = DATA_ALIGN

    # ── Auto-fit column widths ─────────────────────────────────────────────────
    _auto_fit_columns(ws, headers, records, field_names)

    # ── Save ───────────────────────────────────────────────────────────────────
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def _auto_fit_columns(
    ws,
    headers: list[str],
    records: list[dict],
    field_names: list[str],
    min_width: int = 10,
    max_width: int = 60,
) -> None:
    """Set column widths based on header length and data content."""
    for col_idx, (header, field) in enumerate(zip(headers, field_names), start=1):
        # Start with header width
        max_len = len(header)
        for record in records:
            val = _safe_str(record.get(field, ""))
            # For multi-line cells, use the longest line
            cell_max = max((len(line) for line in val.split("\n")), default=0)
            max_len = max(max_len, cell_max)
        col_letter = get_column_letter(col_idx)
        # Add a small padding; clamp to [min_width, max_width]
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


def _safe_str(value: Any) -> str:
    """Convert any value to a string, handling None and nested types."""
    if value is None:
        return ""
    if isinstance(value, list):
        return " | ".join(str(v) for v in value)
    return str(value)
